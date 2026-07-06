## Changes (KeywordScout v2.0 Industry Upgrade — 2026-06-29)
# - [exporter.py]     pandas/pyarrow made optional; CSV/JSON now use stdlib only.
# - [exporter.py]     XLSX uses openpyxl directly (no pandas dependency).

import csv
import io as _io_mod
import json
from typing import Tuple, Optional
from sqlalchemy import func
from sqlalchemy.orm import Session
from backend.models import CrawledURL, SearchQuery

try:
    import pandas as pd
    _PANDAS_AVAILABLE = True
except ImportError:
    pd = None
    _PANDAS_AVAILABLE = False

COLUMN_MAPPING = [
    ("id", "ID"),
    ("search_id", "Search ID"),
    ("url", "URL"),
    ("domain", "Domain"),
    ("title", "Title"),
    ("snippet", "Snippet"),
    ("occurrences", "Occurrences"),
    ("found_in_title", "Found in Title"),
    ("found_in_description", "Found in Description"),
    ("found_in_body", "Found in Body"),
    ("found_in_url", "Found in URL"),
    ("language", "Language"),
    ("status", "Status"),
    ("error_message", "Error Message"),
    ("relevance_score", "Relevance Score"),
    ("is_duplicate", "Is Duplicate"),
    ("content_hash", "Content Hash"),
    ("description", "Description"),
    ("full_content", "Full Content"),
    ("author", "Author"),
    ("image_url", "Image URL"),
    ("image_links", "Image Links"),
    ("video_links", "Video Links"),
    ("discovered_at", "Discovered At"),
    ("matched_keywords", "Matched Keywords"),
]

def get_export_data(
    search_id: int,
    db: Session,
    q: Optional[str] = None,
    status: Optional[str] = None,
    exclude_duplicates: bool = False,
    sort_by: Optional[str] = None,
    sort_desc: bool = True
) -> list:
    """
    Queries matched URLs for a search query and returns them as a list of dicts.
    (Pandas-free — compatible with Python 3.14+)
    """
    query_record = db.query(SearchQuery).filter(SearchQuery.id == search_id).first()
    if not query_record:
        raise ValueError(f"Search Query ID {search_id} not found.")

    query = db.query(CrawledURL).filter(CrawledURL.search_id == search_id)

    if q and q.strip():
        q_clean = q.strip().lower()
        query = query.filter(
            (func.lower(CrawledURL.title).contains(q_clean)) |
            (func.lower(CrawledURL.url).contains(q_clean)) |
            (func.lower(CrawledURL.domain).contains(q_clean))
        )
    if status and status.strip():
        query = query.filter(CrawledURL.status == status.strip())
    if exclude_duplicates:
        query = query.filter(CrawledURL.is_duplicate == False)

    sort_col = CrawledURL.occurrences if sort_by == "occurrences" else CrawledURL.relevance_score
    query = query.order_by(
        sort_col.desc() if sort_desc else sort_col.asc(),
        CrawledURL.discovered_at.desc()
    )

    records = query.all()
    data_list = []
    for r in records:
        matched_kws = ""
        if r.matched_keywords:
            try:
                kws = json.loads(r.matched_keywords)
                matched_kws = ", ".join(kws) if isinstance(kws, list) else str(r.matched_keywords)
            except Exception:
                matched_kws = str(r.matched_keywords)
        data_list.append({
            "ID": r.id,
            "Search ID": r.search_id,
            "URL": r.url,
            "Domain": r.domain,
            "Title": r.title or "",
            "Snippet": r.snippet or "",
            "Occurrences": r.occurrences or 0,
            "Found in Title": bool(r.found_in_title),
            "Found in Description": bool(r.found_in_description),
            "Found in Body": bool(r.found_in_body),
            "Found in URL": bool(r.found_in_url),
            "Language": r.language or "",
            "Status": r.status or "pending",
            "Error Message": r.error_message or "",
            "Relevance Score": float(r.relevance_score or 0.0),
            "Is Duplicate": bool(r.is_duplicate),
            "Content Hash": r.content_hash or "",
            "Description": r.description or "",
            "Full Content": r.full_content or "",
            "Author": r.author or "",
            "Image URL": r.image_url or "",
            "Image Links": r.image_links or "",
            "Video Links": r.video_links or "",
            "Discovered At": r.discovered_at.isoformat() if r.discovered_at else "",
            "Matched Keywords": matched_kws,
        })
    return data_list

def export_results(
    search_id: int,
    format_type: str,
    db: Session,
    q: Optional[str] = None,
    status: Optional[str] = None,
    exclude_duplicates: bool = False,
    sort_by: Optional[str] = None,
    sort_desc: bool = True
) -> Tuple[bytes, str]:
    """
    Exports crawling results in the specified format.
    Returns: (bytes_data, media_type)
    """
    records = get_export_data(search_id, db, q=q, status=status,
                              exclude_duplicates=exclude_duplicates,
                              sort_by=sort_by, sort_desc=sort_desc)
    columns = [title for _, title in COLUMN_MAPPING]
    format_type = format_type.lower()

    if format_type == "csv":
        buf = _io_mod.StringIO()
        writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)
        return buf.getvalue().encode("utf-8-sig"), "text/csv"

    elif format_type == "json":
        return json.dumps(records, ensure_ascii=False, indent=2, default=str).encode("utf-8"), "application/json"

    elif format_type == "xlsx":
        import openpyxl
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook()
        try:
            ws = wb.active
            ws.title = "Keyword Crawl Results"
            skip_cols = {"Occurrences", "Relevance Score", "Content Hash"}
            export_cols = [c for c in columns if c not in skip_cols]
            ws.append(export_cols)
            for row in records:
                ws.append([row.get(c, "") for c in export_cols])
            # Auto-width columns
            for i, col in enumerate(export_cols, 1):
                ws.column_dimensions[get_column_letter(i)].width = min(50, max(12, len(col) + 4))
            buf = _io_mod.BytesIO()
            wb.save(buf)
            buf.seek(0)
            val = buf.getvalue()
        finally:
            # BUGFIX: Ensure openpyxl Workbook is closed to prevent temporary file resource leaks.
            wb.close()
        return val, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    elif format_type == "parquet":
        # Fallback: if pyarrow/pandas available use it; else return JSON with octet-stream
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
            import pandas as _pd
            df = _pd.DataFrame(records, columns=columns)
            buf = _io_mod.BytesIO()
            pq.write_table(pa.Table.from_pandas(df), buf)
            buf.seek(0)
            return buf.getvalue(), "application/octet-stream"
        except ImportError:
            # Graceful degradation: return JSON when pyarrow not available
            return json.dumps(records, ensure_ascii=False, indent=2, default=str).encode("utf-8"), "application/json"

    else:
        raise ValueError(f"Unsupported export format: {format_type}")
